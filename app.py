"""Stine BankReq Reformatter — drag-and-drop the daily raw transaction file and
get back the reformatted BankReq workbook with a new Formatted sheet."""
from __future__ import annotations

import base64
import hashlib
import io
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).parent
LOGO_FILE = ROOT / "Stinelogo_white_rec.svg"

st.set_page_config(page_title="Stine CardConnect Reformatter", layout="wide")

st.markdown(
    """
    <style>
    header[data-testid="stHeader"] {display: none !important;}
    [data-testid="stToolbar"] {display: none !important;}
    .stDeployButton {display: none !important;}
    [data-testid="stDeployButton"] {display: none !important;}
    [data-testid="stAppDeployButton"] {display: none !important;}
    </style>
    """,
    unsafe_allow_html=True,
)

OUT_COLS = [
    "Site Alternate ID",
    "Funded Date",
    "Site Name",
    "Product Code",
    "Processed Transaction Amount",
]

PRODUCT_ORDER = ["Amex", "DebitCard", "Discover", "Mastercard", "Visa"]

NET_SALES_SHEET = "Net Sales"
ADJUSTMENTS_SHEET = "Adjustments"
CHARGEBACKS_SHEET = "Chargebacks & Chargeback Revers"
RED = "FFFF0000"

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

THICK = Side(style="thick")
THIN = Side(style="thin")
BOLD = Font(bold=True)
RED_FONT = Font(color=RED)


def _product_key(p: object) -> tuple[int, int, str]:
    s = str(p)
    if s in PRODUCT_ORDER:
        return (0, PRODUCT_ORDER.index(s), s)
    return (1, 0, s)


def _funded_date_str(value: object) -> str:
    if isinstance(value, str):
        return value
    if pd.isna(value):
        return ""
    try:
        return pd.to_datetime(value).strftime("%m/%d/%Y")
    except Exception:
        return str(value)


def _site_key(value: object) -> tuple[int, float, str]:
    s = str(value)
    try:
        return (0, float(s), s)
    except (TypeError, ValueError):
        return (1, 0.0, s)


def _group_section(df: pd.DataFrame) -> pd.DataFrame:
    needed = {
        "Site Alternate ID",
        "Funded Date",
        "Site Name",
        "Product Code",
        "Processed Transaction Amount",
    }
    if not needed.issubset(df.columns):
        return pd.DataFrame(columns=OUT_COLS)

    work = df[list(needed)].copy()
    work["Funded Date"] = work["Funded Date"].map(_funded_date_str)
    grouped = (
        work.groupby(
            ["Site Alternate ID", "Funded Date", "Site Name", "Product Code"],
            dropna=False,
            as_index=False,
        )["Processed Transaction Amount"]
        .sum()
    )
    grouped["__site"] = grouped["Site Alternate ID"].map(_site_key)
    grouped["__prod"] = grouped["Product Code"].map(_product_key)
    grouped = grouped.sort_values(["__site", "__prod"]).drop(columns=["__site", "__prod"])
    return grouped[OUT_COLS].reset_index(drop=True)


def build_branches(sheets: dict[str, pd.DataFrame]) -> list[dict]:
    """Merge every site's Net Sales, Adjustments, and Chargebacks into a single
    branch dict. Adjustments and chargeback rows are tagged red so the writer
    can render them in red text."""
    sources = [
        (NET_SALES_SHEET, "net"),
        (ADJUSTMENTS_SHEET, "adj"),
        (CHARGEBACKS_SHEET, "cb"),
    ]
    by_branch: dict[tuple, dict] = {}
    for sheet_name, role in sources:
        df = sheets.get(sheet_name)
        if df is None or df.empty:
            continue
        block = _group_section(df)
        for _, row in block.iterrows():
            key = (str(row["Site Alternate ID"]), row["Funded Date"], row["Site Name"])
            if key not in by_branch:
                by_branch[key] = {
                    "Site Alternate ID": key[0],
                    "Funded Date": key[1],
                    "Site Name": key[2],
                    "net": [],
                    "adj": [],
                    "cb": [],
                }
            by_branch[key][role].append(
                (row["Product Code"], row["Processed Transaction Amount"])
            )
    return sorted(by_branch.values(), key=lambda b: _site_key(b["Site Alternate ID"]))


def derive_output_filename(branches: list[dict]) -> str:
    for branch in branches:
        raw = branch.get("Funded Date")
        if not raw:
            continue
        try:
            dt = pd.to_datetime(raw)
            return f"BankReq - {dt.strftime('%m-%d-%Y')}.xlsx"
        except Exception:
            break
    return f"BankReq - {datetime.now().strftime('%m-%d-%Y')}.xlsx"


def _apply_border(cell, *, left=None, right=None, top=None, bottom=None) -> None:
    b = cell.border
    cell.border = Border(
        left=left if left is not None else b.left,
        right=right if right is not None else b.right,
        top=top if top is not None else b.top,
        bottom=bottom if bottom is not None else b.bottom,
    )


def _draw_thick_box(ws: Worksheet, top_row: int, bottom_row: int, left_col: int = 1, right_col: int = 6) -> None:
    for c in range(left_col, right_col + 1):
        _apply_border(ws.cell(row=top_row, column=c), top=THICK)
        _apply_border(ws.cell(row=bottom_row, column=c), bottom=THICK)
    for r in range(top_row, bottom_row + 1):
        _apply_border(ws.cell(row=r, column=left_col), left=THICK)
        _apply_border(ws.cell(row=r, column=right_col), right=THICK)


def _write_data_row(
    ws: Worksheet,
    row: int,
    branch: dict,
    product: str,
    amount: object,
    *,
    red: bool = False,
) -> None:
    site_id = branch["Site Alternate ID"]
    cells = [
        ws.cell(row=row, column=1, value=str(site_id) if site_id is not None else None),
        ws.cell(row=row, column=2, value=branch["Funded Date"]),
        ws.cell(row=row, column=3, value=branch["Site Name"]),
        ws.cell(row=row, column=4, value=product),
        ws.cell(
            row=row,
            column=5,
            value=float(amount) if pd.notna(amount) else None,
        ),
    ]
    if red:
        for c in cells:
            c.font = RED_FONT


def _write_branch(ws: Worksheet, start_row: int, branch: dict) -> int:
    net_rows = branch["net"]
    amex = [(p, a) for p, a in net_rows if str(p) == "Amex"]
    others = [(p, a) for p, a in net_rows if str(p) != "Amex"]
    others.sort(key=lambda x: _product_key(x[0]))

    cur = start_row
    for product, amount in amex:
        _write_data_row(ws, cur, branch, product, amount)
        cur += 1

    box_top = cur
    for product, amount in others:
        _write_data_row(ws, cur, branch, product, amount)
        cur += 1
    for product, amount in branch["adj"]:
        _write_data_row(ws, cur, branch, product, amount, red=True)
        cur += 1
    for product, amount in branch["cb"]:
        _write_data_row(ws, cur, branch, product, amount, red=True)
        cur += 1

    if cur > box_top:
        box_bottom = cur - 1
        _draw_thick_box(ws, box_top, box_bottom)
        ws.cell(
            row=box_bottom,
            column=6,
            value=f"=SUM(E{box_top}:E{box_bottom})",
        )

    return cur


def write_workbook(raw_bytes: bytes, branches: list[dict]) -> bytes:
    wb = load_workbook(io.BytesIO(raw_bytes))
    if "Formatted" in wb.sheetnames:
        del wb["Formatted"]
    ws = wb.create_sheet("Formatted", 0)

    for col_idx, name in enumerate(OUT_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = BOLD
        _apply_border(cell, bottom=THIN)

    cur = 2
    for i, branch in enumerate(branches):
        if i > 0:
            cur += 1  # blank row between branches for readability
        cur = _write_branch(ws, cur, branch)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def trigger_browser_download(data: bytes, filename: str, token: str) -> None:
    if st.session_state.get("_last_auto_download") == token:
        return
    st.session_state["_last_auto_download"] = token
    b64 = base64.b64encode(data).decode("ascii")
    components.html(
        f"""
        <script>
          const a = document.createElement('a');
          a.href = "data:{XLSX_MIME};base64,{b64}";
          a.download = {filename!r};
          document.body.appendChild(a);
          a.click();
          a.remove();
        </script>
        """,
        height=0,
    )


def main() -> None:
    col_logo, col_title = st.columns([1, 5], vertical_alignment="center")
    with col_logo:
        if LOGO_FILE.exists():
            st.image(str(LOGO_FILE), width=160)
    with col_title:
        st.title("CardConnect Reformatter")
        st.caption(
            "Drop the daily raw credit-card transaction workbook below. "
            "The formatted BankReq file will download automatically."
        )

    uploaded = st.file_uploader(
        "Drag and drop the raw .xlsx file here, or click to browse",
        type=["xlsx"],
        accept_multiple_files=False,
    )
    if uploaded is None:
        return

    raw_bytes = uploaded.getvalue()
    try:
        sheets = pd.read_excel(io.BytesIO(raw_bytes), sheet_name=None)
    except Exception as e:
        st.error(f"Could not read the uploaded workbook: {e}")
        return

    branches = build_branches(sheets)
    if not branches:
        st.error(
            "No usable rows found. Expected a 'Net Sales' sheet with the "
            "standard BankReq columns."
        )
        return

    output_bytes = write_workbook(raw_bytes, branches)
    filename = derive_output_filename(branches)

    token = hashlib.sha1(raw_bytes).hexdigest()
    trigger_browser_download(output_bytes, filename, token)

    st.success(f"Formatted file ready: **{filename}**")
    st.download_button(
        "Download again",
        data=output_bytes,
        file_name=filename,
        mime=XLSX_MIME,
    )

    preview_rows = []
    for branch in branches:
        for product, amount in branch["net"]:
            preview_rows.append(
                {
                    "Site Alternate ID": branch["Site Alternate ID"],
                    "Funded Date": branch["Funded Date"],
                    "Site Name": branch["Site Name"],
                    "Product Code": product,
                    "Processed Transaction Amount": amount,
                    "Source": "Net Sales",
                }
            )
        for product, amount in branch["adj"]:
            preview_rows.append(
                {
                    "Site Alternate ID": branch["Site Alternate ID"],
                    "Funded Date": branch["Funded Date"],
                    "Site Name": branch["Site Name"],
                    "Product Code": product,
                    "Processed Transaction Amount": amount,
                    "Source": "Adjustments",
                }
            )
        for product, amount in branch["cb"]:
            preview_rows.append(
                {
                    "Site Alternate ID": branch["Site Alternate ID"],
                    "Funded Date": branch["Funded Date"],
                    "Site Name": branch["Site Name"],
                    "Product Code": product,
                    "Processed Transaction Amount": amount,
                    "Source": "Chargebacks",
                }
            )
    if preview_rows:
        st.subheader("Preview")
        preview_df = pd.DataFrame(preview_rows)

        def _highlight_red(row):
            return [
                "color: red" if row["Source"] in ("Adjustments", "Chargebacks") else ""
                for _ in row
            ]

        st.dataframe(
            preview_df.style.apply(_highlight_red, axis=1),
            use_container_width=True,
            hide_index=True,
        )


if __name__ == "__main__":
    main()
