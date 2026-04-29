from collections import Counter, defaultdict
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.workbook import Workbook

from . import history

SOURCE_SHEET = "Net Sales"
ADDITIONAL_SOURCE_SHEETS = ("Adjustments", "Chargebacks & Chargeback Revers")
OUTPUT_SHEET_BASE = "Formatted"

OUTPUT_HEADERS = [
    "Site Alternate ID",
    "Funded Date",
    "Site Name",
    "Product Code",
    "Processed Transaction Amount",
    None,
]

AMOUNT_FORMAT = "#,###.00;\\-#,###.00"

# Anomaly highlight: light orange fill, matches the Stine accent palette.
ANOMALY_FILL = PatternFill(
    fill_type="solid", start_color="FDEEE5", end_color="FDEEE5"
)
ANOMALY_K = 2.0  # how many stddevs out is "anomalous"

COLUMN_WIDTHS = {
    "A": 13.0,
    "B": 12.3,
    "C": 22.7,
    "D": 14.0,
    "E": 17.4,
    "F": 14.6,
}

THIN = Side(style="thin")
THICK = Side(style="thick")
NONE_SIDE = Side(style=None)


def _product_sort_key(pc: str) -> str:
    return (pc or "").lower()


REQUIRED_COLUMNS = (
    "Site Alternate ID",
    "Site Name",
    "Funded Date",
    "Product Code",
    "Processed Transaction Amount",
)


def _ingest_sheet(ws, aggregated, site_meta, date_counts) -> bool:
    """Read rows from a single sheet into the shared aggregates. Returns
    True if the sheet had a recognisable header, False otherwise (which
    we treat as 'empty/blank' and skip silently)."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return False
    headers = list(header_row)
    if not set(REQUIRED_COLUMNS).issubset(set(headers)):
        return False
    idx = {name: headers.index(name) for name in REQUIRED_COLUMNS}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(v is None for v in row):
            continue
        alt_id = row[idx["Site Alternate ID"]]
        sname = row[idx["Site Name"]]
        fdate = row[idx["Funded Date"]]
        pc = row[idx["Product Code"]]
        amt = row[idx["Processed Transaction Amount"]]
        if alt_id is None or pc is None or amt is None:
            continue
        aggregated[(alt_id, pc)] += float(amt)
        site_meta.setdefault(alt_id, (sname, fdate))
        if fdate is not None:
            date_counts[fdate] += 1
    return True


def _read_source_rows(wb: Workbook):
    """Return (sources, site_meta, funded_date) where sources is a dict
    keyed by source-sheet name -> {(alt_id, product_code): amount}.
    Net Sales is required; the others are optional and skipped if absent
    or empty."""
    if SOURCE_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Workbook is missing the required '{SOURCE_SHEET}' sheet. "
            f"Found sheets: {wb.sheetnames}"
        )

    sources: dict[str, dict[tuple, float]] = {}
    site_meta: dict[str, tuple[str, str]] = {}
    date_counts: Counter = Counter()

    net_sales: dict[tuple, float] = defaultdict(float)
    if not _ingest_sheet(wb[SOURCE_SHEET], net_sales, site_meta, date_counts):
        raise ValueError(
            f"'{SOURCE_SHEET}' sheet is missing required columns: "
            f"{sorted(REQUIRED_COLUMNS)}"
        )
    sources[SOURCE_SHEET] = net_sales

    for sheet_name in ADDITIONAL_SOURCE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        agg: dict[tuple, float] = defaultdict(float)
        _ingest_sheet(wb[sheet_name], agg, site_meta, date_counts)
        if agg:
            sources[sheet_name] = agg

    funded_date = date_counts.most_common(1)[0][0] if date_counts else None
    return sources, site_meta, funded_date


def _next_sheet_name(wb: Workbook) -> str:
    if OUTPUT_SHEET_BASE not in wb.sheetnames:
        return OUTPUT_SHEET_BASE
    n = 2
    while f"{OUTPUT_SHEET_BASE} ({n})" in wb.sheetnames:
        n += 1
    return f"{OUTPUT_SHEET_BASE} ({n})"


def _box_border(top: bool, bottom: bool, left: bool, right: bool) -> Border:
    return Border(
        top=THICK if top else NONE_SIDE,
        bottom=THICK if bottom else NONE_SIDE,
        left=THICK if left else NONE_SIDE,
        right=THICK if right else NONE_SIDE,
    )


def _draw_box(ws, top_row: int, bottom_row: int) -> None:
    for r in range(top_row, bottom_row + 1):
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = _box_border(
                top=(r == top_row),
                bottom=(r == bottom_row),
                left=(c == 1),
                right=(c == 6),
            )


def _flag_anomaly(amt_cell, alt_id, pc, source_sheet, amount, asof) -> bool:
    """Apply an anomaly highlight to amt_cell if `amount` is an outlier
    compared to history for (alt_id, pc, source_sheet) before `asof`.
    Returns True when a flag was applied."""
    if asof is None:
        return False
    bl = history.baseline(alt_id, pc, source_sheet, asof)
    if bl is None:
        return False
    lo, hi = bl.lower(ANOMALY_K), bl.upper(ANOMALY_K)
    if lo <= amount <= hi:
        return False
    amt_cell.fill = ANOMALY_FILL
    direction = "above" if amount > hi else "below"
    amt_cell.comment = Comment(
        (
            f"Anomaly: {amount:,.2f} is {direction} the typical range.\n"
            f"Last {bl.n} day(s): mean={bl.mean:,.2f}, "
            f"normal range={lo:,.2f} – {hi:,.2f}."
        ),
        "Bank Reformatter",
    )
    return True


def _write_net_sales_section(
    ws, aggregated, site_meta, start_row: int, asof, anomaly_counter
) -> int:
    """Write the per-site Net Sales boxes. Returns the next free row."""
    sites = sorted(site_meta.keys(), key=lambda s: str(s))
    current_row = start_row
    for alt_id in sites:
        sname, fdate = site_meta[alt_id]
        products = sorted(
            [pc for (a, pc) in aggregated.keys() if a == alt_id],
            key=_product_sort_key,
        )
        if not products:
            continue
        site_start_row = current_row
        first_non_amex_row: int | None = None
        for pc in products:
            amt = aggregated[(alt_id, pc)]
            ws.cell(row=current_row, column=1, value=alt_id)
            ws.cell(row=current_row, column=2, value=fdate)
            ws.cell(row=current_row, column=3, value=sname)
            ws.cell(row=current_row, column=4, value=pc)
            amt_cell = ws.cell(row=current_row, column=5, value=amt)
            amt_cell.number_format = AMOUNT_FORMAT
            if _flag_anomaly(amt_cell, alt_id, pc, SOURCE_SHEET, amt, asof):
                anomaly_counter[0] += 1
            if pc != "Amex" and first_non_amex_row is None:
                first_non_amex_row = current_row
            current_row += 1
        last_row = current_row - 1

        if last_row > site_start_row and first_non_amex_row is not None:
            sub_cell = ws.cell(row=last_row, column=6)
            sub_cell.value = f"=SUM(E{first_non_amex_row}:E{last_row})"
            sub_cell.number_format = AMOUNT_FORMAT

        box_start = first_non_amex_row if first_non_amex_row else site_start_row
        if box_start is not None and box_start <= last_row:
            _draw_box(ws, box_start, last_row)
    return current_row


def _write_aux_section(
    ws, label: str, agg: dict, site_meta: dict, start_row: int,
    asof, anomaly_counter,
) -> int:
    """Write a labeled, boxed section for an auxiliary data source
    (Adjustments, Chargebacks). Returns the next free row."""
    label_row = start_row
    ws.merge_cells(
        start_row=label_row, start_column=1, end_row=label_row, end_column=6
    )
    label_cell = ws.cell(row=label_row, column=1, value=f"From: {label}")
    label_cell.font = Font(bold=True, color="FFFFFF", size=11)
    label_cell.fill = PatternFill(
        fill_type="solid", start_color="008445", end_color="008445"
    )

    data_start = label_row + 1
    current_row = data_start
    keys = sorted(agg.keys(), key=lambda k: (str(k[0]), _product_sort_key(k[1])))
    for alt_id, pc in keys:
        sname, fdate = site_meta.get(alt_id, (None, None))
        amt = agg[(alt_id, pc)]
        ws.cell(row=current_row, column=1, value=alt_id)
        ws.cell(row=current_row, column=2, value=fdate)
        ws.cell(row=current_row, column=3, value=sname)
        ws.cell(row=current_row, column=4, value=pc)
        amt_cell = ws.cell(row=current_row, column=5, value=amt)
        amt_cell.number_format = AMOUNT_FORMAT
        if _flag_anomaly(amt_cell, alt_id, pc, label, amt, asof):
            anomaly_counter[0] += 1
        current_row += 1

    data_last = current_row - 1
    if data_last >= data_start:
        total_cell = ws.cell(row=data_last, column=6)
        total_cell.value = f"=SUM(E{data_start}:E{data_last})"
        total_cell.number_format = AMOUNT_FORMAT
        _draw_box(ws, label_row, data_last)
    else:
        _draw_box(ws, label_row, label_row)

    return current_row


def _write_output_sheet(wb: Workbook, sources, site_meta, asof) -> tuple[str, int]:
    name = _next_sheet_name(wb)
    ws = wb.create_sheet(name, 0)
    wb.active = 0

    bold = Font(bold=True)
    header_border = Border(bottom=THIN)
    for col_idx, header in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if header is not None:
            cell.font = bold
            cell.border = header_border

    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    anomaly_counter = [0]
    next_row = _write_net_sales_section(
        ws, sources.get(SOURCE_SHEET, {}), site_meta,
        start_row=2, asof=asof, anomaly_counter=anomaly_counter,
    )

    for sheet_name in ADDITIONAL_SOURCE_SHEETS:
        agg = sources.get(sheet_name)
        if not agg:
            continue
        next_row += 1  # blank separator row
        next_row = _write_aux_section(
            ws, sheet_name, agg, site_meta,
            start_row=next_row, asof=asof, anomaly_counter=anomaly_counter,
        )

    return name, anomaly_counter[0]


def _normalize_date(raw) -> date:
    """Return a date object from the funded-date cell value, falling back to
    today if it can't be parsed."""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    return date.today()


def reformat_workbook(file_bytes: bytes) -> tuple[bytes, str, date, int]:
    """Take a raw merchant-services workbook, append a 'Formatted' sheet
    (auto-suffixed if one already exists), and return
    (workbook_bytes, sheet_name, funded_date, anomaly_count).

    Anomaly highlighting compares each (site, product, source) cell to
    the trailing 60-day baseline persisted in the local history DB. The
    current upload is recorded after we render, so its values become
    part of the baseline for *future* uploads."""
    wb = load_workbook(BytesIO(file_bytes))
    sources, site_meta, raw_date = _read_source_rows(wb)
    funded_date = _normalize_date(raw_date)
    sheet_name, anomaly_count = _write_output_sheet(
        wb, sources, site_meta, asof=funded_date
    )
    # Persist this upload to history so future runs have a baseline.
    history_rows = []
    for source_name, agg in sources.items():
        for (alt_id, pc), amt in agg.items():
            sname = site_meta.get(alt_id, (None, None))[0]
            history_rows.append((alt_id, sname, pc, source_name, amt))
    history.record_upload(history_rows, funded_date=funded_date)

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), sheet_name, funded_date, anomaly_count
